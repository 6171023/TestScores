import openpyxl
from django.shortcuts import render
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.http import HttpResponse, HttpResponseNotFound
from fuzzywuzzy import fuzz
from collections import OrderedDict
from .forms import UploadFileForm
import os
from openpyxl.utils import column_index_from_string

def download_file(request, file_name):
    file_path = default_storage.path(file_name)
    if default_storage.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    return HttpResponseNotFound("File not found.")

def success(request, file_name):
    return render(request, 'success.html', {'modified_file_name': file_name})

def upload(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            overwrite_file = request.FILES['overwrite_file']
            extract_file = request.FILES['extract_file']
            column_to_overwrite = form.cleaned_data['column_to_overwrite']  # Get the column to overwrite from the form

            if overwrite_file.name == extract_file.name:
                form.add_error(None, "The names of the uploaded files cannot be the same.")
                return render(request, 'home.html', {'form': form})
            
            overwrite_path = default_storage.save(overwrite_file.name, ContentFile(overwrite_file.read()))
            extract_path = default_storage.save(extract_file.name, ContentFile(extract_file.read()))

            overwrite_data = process_overwrite_file(default_storage.path(overwrite_path))
            extract_data = process_extract_file(default_storage.path(extract_path))
            
            separated_batches = separate_emails_into_batches(extract_data, overwrite_data)
            
            modified_file_path = modify_overwrite_file(default_storage.path(overwrite_path), separated_batches, column_to_overwrite)
            
            with open(modified_file_path, 'rb') as f:
                modified_file = ContentFile(f.read())
                modified_file_name = default_storage.save('modified_file.xlsx', modified_file)

            default_storage.delete(overwrite_path)
            default_storage.delete(extract_path)
            
            if os.path.exists(modified_file_path):
                os.remove(modified_file_path)

            download_link = default_storage.url(modified_file_name)

            return render(request, 'success.html', {'download_link': download_link})
    else:
        form = UploadFileForm()
    
    return render(request, 'upload.html', {'form': form})

def process_overwrite_file(file_path):
    # Open the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Attendance']

    # Extract the data
    batch_data = {}
    current_batch = []
    empty_cells_count = 0
    batch_index = 1

    for row in sheet.iter_rows(min_row=6, min_col=8, max_col=8):
        email_cell_value = row[0].value
        full_name_cell_value = sheet.cell(row=row[0].row, column=2).value  # Full name from Column 2

        if email_cell_value:
            current_batch.append((email_cell_value.replace(' ', ''), full_name_cell_value))  # Remove spaces from email
            empty_cells_count = 0  # Reset the empty cell counter when data is found
        else:
            empty_cells_count += 1
            if empty_cells_count > 2:
                break  # Stop reading after more than 2 empty cells
            if empty_cells_count == 1 and current_batch:
                batch_name = f"batch_{batch_index}"
                batch_data[batch_name] = current_batch
                current_batch = []
                batch_index += 1

    if current_batch:
        batch_name = f"batch_{batch_index}"
        batch_data[batch_name] = current_batch

    #for debug
    # print(batch_data)
    # total_values = sum(len(batch) for batch in batch_data.values())
    # print(f"Total number of values in all batches: {total_values}")
    return batch_data

def process_extract_file(file_path):
    # Open the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Test Takers']

    # Extract the emails, full names, and scores
    data = []

    empty_rows_count = 0
    for row in sheet.iter_rows(min_row=5, min_col=3, max_col=7):
        email = row[0].value
        full_name = sheet.cell(row=row[0].row, column=2).value  # Full name from Column 2
        score = row[4].value
        if email and score is not None:
            data.append((email.replace(' ', ''), full_name, score))  # Remove spaces from emails
            empty_rows_count = 0  # Reset the empty row counter when data is found
        elif not email:
            empty_rows_count += 1
            if empty_rows_count > 2:
                break  # Stop reading after more than 2 empty rows

    return data

def separate_emails_into_batches(data, batch_data):
    separated_batches = OrderedDict()

    # Initialize separated_batches with batch names in the order they appear in batch_data
    for batch_name in batch_data:
        separated_batches[batch_name] = []

    # Iterate over batch_data to process batches in order
    for batch_name, batch_emails in batch_data.items():
        matched_data = []  # Temporary list to collect matched data

        # Iterate over emails and check against batch_emails using fuzzy matching
        for batch_email, batch_full_name in batch_emails:
            found = False
            for email, full_name, score in data:
                if fuzz.ratio(email.lower(), batch_email.lower()) >= 90:
                    matched_data.append((batch_email, batch_full_name, score))  # Use batch_email and batch_full_name from batch_data
                    found = True
                    break  # Move to the next batch_email once a match is found
            if not found:
                matched_data.append((batch_email, batch_full_name, None))  # Use batch_email and batch_full_name from batch_data even if no match is found

        # Extend matched data to the corresponding batch in the correct order
        separated_batches[batch_name].extend(matched_data)

    # print(f"Length of separated_batches: {len(separated_batches)}")
    return separated_batches

def modify_overwrite_file(file_path, separated_batches, column_to_overwrite):
    # Load the original workbook
    wb = openpyxl.load_workbook(file_path)
    modified_wb = openpyxl.Workbook()

    # Copy the sheets from the original workbook to the new workbook
    for sheet_name in wb.sheetnames:
        original_sheet = wb[sheet_name]
        modified_sheet = modified_wb.create_sheet(title=sheet_name)

        for row in original_sheet.iter_rows():
            for cell in row:
                modified_sheet[cell.coordinate].value = cell.value

    # Remove the default sheet created by openpyxl
    del modified_wb['Sheet']

    # Modify the 'Test Scores' sheet
    test_scores_sheet = modified_wb['Test Scores']

    # Convert column letter to column index
    column_index = openpyxl.utils.column_index_from_string(column_to_overwrite)

    # Iterate over the rows and match the full names to update scores
    for row in test_scores_sheet.iter_rows(min_row=7):
        full_name = row[1].value  # Column 2 is the full name
        if full_name:  # Ensure full_name is not None
            for batch_name, batch_data in separated_batches.items():
                for email, batch_full_name, score in batch_data:
                    if batch_full_name and full_name.strip().lower() == batch_full_name.strip().lower():
                        if score is not None:  # Ensure score is not None
                            row[column_index - 1].value = score  # Update the score in the sheet
                        else:
                            row[column_index - 1].value = None  # Clear the cell if score is None
                        break

    # Save the modified workbook to a file
    modified_file_path = 'modified_file.xlsx'
    modified_wb.save(modified_file_path)

    return modified_file_path

